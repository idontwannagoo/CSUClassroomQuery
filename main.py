import openpyxl
import re
import json
from itertools import islice
from tqdm import tqdm
import time
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def parse_week_numbers(week_text):
    """解析周数文本，返回包含所有周数的集合"""
    weeks = set()
    
    # 判断是否有单双周限定
    is_odd = '单周' in week_text
    is_even = '双周' in week_text
    
    # 移除"周"字和单双周标记，便于解析
    week_text = week_text.replace('周', '')
    # 注意：不要移除"单"和"双"字，因为它们可能出现在中间
    
    # 分割多个部分（逗号分隔）
    parts = week_text.split(',')
    
    for part in parts:
        # 检查这部分是否有单双周限定
        part_is_odd = is_odd or '单' in part
        part_is_even = is_even or '双' in part
        
        # 移除可能的单双周标记
        part = part.replace('单', '').replace('双', '')
        
        if '-' in part:  # 处理区间，如"1-8"
            start, end = map(int, part.split('-'))
            if part_is_odd:
                weeks.update(w for w in range(start, end + 1) if w % 2 == 1)
            elif part_is_even:
                weeks.update(w for w in range(start, end + 1) if w % 2 == 0)
            else:
                weeks.update(range(start, end + 1))
        else:  # 处理单个数字，如"9"
            week_num = int(part)
            if (not part_is_odd and not part_is_even) or \
               (part_is_odd and week_num % 2 == 1) or \
               (part_is_even and week_num % 2 == 0):
                weeks.add(week_num)
    
    return weeks

def parse_classroom_info(text):
    if not isinstance(text, str):
        return None
        
    # 匹配教室信息的正则表达式
    classroom_pattern = r'[A-D]座\d+'
    
    # 匹配周数信息的正则表达式（更复杂的模式）
    week_pattern = r'(\d+(?:-\d+)?(?:,\d+(?:-\d+)?)*(?:单|双)?周)'
    
    # 查找教室信息
    classroom = re.search(classroom_pattern, text)
    if not classroom:
        return None
    
    # 查找周数信息
    week_info = re.search(week_pattern, text)
    if not week_info:
        return None
    
    # 解析周数
    weeks = parse_week_numbers(week_info.group(1))
    if not weeks:
        return None
    
    return {
        'weeks': weeks,
        'classroom': classroom.group()
    }

def merge_classroom_results(results):
    """合并教室使用信息，去除重复并组合相同教室的不同时段"""
    # 用字典存储每个教室的周数信息
    classroom_weeks = {}
    
    for result in results:
        # 分割周数和教室信息
        weeks_part, classroom = result.split(' ', 1)
        weeks_part = weeks_part.replace('第', '').replace('周', '')
        
        # 将周数字符串转换为集合
        weeks = set(int(w) for w in weeks_part.split(','))
        
        # 更新字典
        if classroom in classroom_weeks:
            classroom_weeks[classroom].update(weeks)
        else:
            classroom_weeks[classroom] = weeks
    
    # 将合并后的信息转换回格式化的字符串
    merged_results = []
    for classroom, weeks in classroom_weeks.items():
        # 将周数排序并转换为连续区间
        weeks_list = sorted(list(weeks))
        intervals = []
        start = end = weeks_list[0]
        
        for week in weeks_list[1:]:
            if week == end + 1:
                end = week
            else:
                intervals.append(f"{start}-{end}" if start != end else str(start))
                start = end = week
        intervals.append(f"{start}-{end}" if start != end else str(start))
        
        # 生成最终的字符串
        weeks_str = ','.join(intervals)
        merged_results.append(f"第{weeks_str}周 {classroom}")
    
    return sorted(merged_results)

def read_excel_cells(filename, base_row, col):
    try:
        print("\n正在读取Excel文件...")
        # 使用只读模式打开Excel文件，并禁用样式加载
        workbook = openpyxl.load_workbook(
            filename, 
            read_only=True, 
            data_only=True,
            keep_links=False
        )
        sheet = workbook.active
        
        results = []
        classroom_infos = []
        
        # 使用迭代器获取指定列的所有单元格
        column_cells = list(sheet.iter_rows(
            min_row=base_row,
            max_row=None,  # 读取到最后一行
            min_col=col,
            max_col=col,
            values_only=True
        ))
        
        # 创建进度条
        total_cells = len(column_cells)
        with tqdm(total=total_cells//10 + 1, desc="处理数据", ncols=100) as pbar:
            # 每隔10行取一个值
            for cell in islice(column_cells, 0, None, 10):
                cell_value = cell[0]  # 因为只取了一列，所以是第一个元素
                
                if cell_value:
                    if isinstance(cell_value, str):
                        cell_value = cell_value.replace('_x000D_', '')
                        # 处理多行课程信息
                        for course in cell_value.split('\n'):
                            info = parse_classroom_info(course)
                            if info:
                                # 格式化周数显示
                                weeks_list = sorted(info['weeks'])
                                weeks_str = ','.join(str(w) for w in weeks_list)
                                results.append(
                                    f"第{weeks_str}周 {info['classroom']}"
                                )
                                classroom_infos.append(info)
                
                pbar.update(1)
        
        # 合并和去重结果
        merged_results = merge_classroom_results(results)
        
        workbook.close()
        return merged_results, classroom_infos if merged_results else (["没有找到符合条件的教室信息"], [])
    
    except FileNotFoundError:
        return [f"错误：找不到文件 '{filename}'"], []
    except Exception as e:
        return [f"错误：{str(e)}"], []

def get_classrooms_by_week(classroom_infos, week):
    """获取指定周的所有教室（去重）"""
    return sorted(set(
        info['classroom'] 
        for info in classroom_infos 
        if week in info['weeks']
    ))

def load_classroom_info():
    """加载教室信息"""
    try:
        print("正在加载教室配置...")
        with open('classrooms.json', 'r', encoding='utf-8') as f:
            data = json.load(f)
            # 添加加载进度条
            result = {}
            for room in tqdm(data, desc="加载教室信息", ncols=100):
                result[room['room_number']] = room
                # time.sleep(0.01)  # 添加小延迟使进度条更容易观察
            return result
    except FileNotFoundError:
        print("警告：找不到classrooms.json文件")
        return {}
    except json.JSONDecodeError:
        print("警告：classrooms.json文件格式错误")
        return {}

def get_available_classrooms(used_classrooms, all_classrooms):
    """获取可用教室列表"""
    available = []
    for room_number, room_info in all_classrooms.items():
        if room_number not in used_classrooms:
            available.append({
                'room_number': room_number,
                'room_type': room_info.get('room_type', '未知类型'),
                'capacity': room_info.get('capacity', '未知容量')
            })
    return sorted(available, key=lambda x: x['room_number'])

def get_row_col_by_time(day, period):
    """
    将周几和节次转换为对应的行列数
    day: 1-5 (周一到周五)
    period: 1-5 (1: 1-2节, 2: 3-4节, 3: 5-6节, 4: 7-8节, 5: 9-10节)
    返回: (row, col)
    """
    # 基准行数计算：每个时间段对应的基准行
    period_to_row = {
        1: 4,   # 1-2节
        2: 5,   # 3-4节
        3: 6,   # 5-6节
        4: 7,   # 7-8节
        5: 8    # 9-10节
    }
    
    # 列数计算：周一从第2列开始
    col = day + 1
    
    # 获取对应的行数
    row = period_to_row.get(period)
    
    if not row:
        raise ValueError("无效的节次")
    
    return row, col

def format_time_period(period):
    """将节次数转换为具体的节次范围"""
    period_map = {
        1: "1-2",
        2: "3-4",
        3: "5-6",
        4: "7-8",
        5: "9-10"
    }
    return period_map.get(period, "未知")

def export_to_excel(all_classrooms):
    """
    自动查询并导出空闲教室到Excel
    """
    print("\n开始自动查询并导出空闲教室信息...")
    
    # 创建新的Excel工作簿
    output_workbook = openpyxl.Workbook()
    
    # 设置进度条总数（5天 * 5节 = 25个时间段）
    total_queries = 5 * 5
    with tqdm(total=total_queries, desc="总体进度", ncols=100) as main_pbar:
        # 先查询所有时间段的课程信息
        time_slots = {}  # 存储所有时间段的查询结果
        
        # 查询每个时间段
        for day in range(1, 6):  # 周一到周五
            for period in range(1, 6):  # 5个节次
                # 获取对应的行列
                row, col = get_row_col_by_time(day, period)
                
                # 查询该时段的教室使用情况
                results, classroom_infos = read_excel_cells('input1.xlsx', row, col)
                # 确保classroom_infos是列表而不是字符串
                if isinstance(classroom_infos, list):
                    time_slots[(day, period)] = classroom_infos
                else:
                    time_slots[(day, period)] = []
                
                # 更新进度条
                main_pbar.update(1)
        
        # 为每周创建工作表
        print("\n正在生成Excel表格...")
        for week in range(1, 17):
            # 创建该周的工作表
            sheet_name = f"第{week}周"
            if week == 1:
                sheet = output_workbook.active
                sheet.title = sheet_name
            else:
                sheet = output_workbook.create_sheet(sheet_name)
            
            # 设置表头
            days = ["", "周一", "周二", "周三", "周四", "周五"]
            for col, day in enumerate(days, 1):
                sheet.cell(row=1, column=col, value=day)
            
            periods = ["", "1-2节", "3-4节", "5-6节", "7-8节", "9-10节"]
            for row, period in enumerate(periods, 1):
                sheet.cell(row=row, column=1, value=period)
            
            # 填充数据
            for day in range(1, 6):
                for period in range(1, 6):
                    classroom_infos = time_slots[(day, period)]
                    
                    try:
                        # 获取已使用的教室
                        used_classrooms = get_classrooms_by_week(classroom_infos, week)
                        
                        # 获取可用教室
                        available = get_available_classrooms(used_classrooms, all_classrooms)
                        
                        # 将可用教室号写入单元格
                        cell_value = " ".join(room['room_number'] for room in available)
                        sheet.cell(row=period+1, column=day+1, value=cell_value)
                    except Exception as e:
                        print(f"处理数据时出错 (周{week}, 星期{day}, 第{period}节): {str(e)}")
                        sheet.cell(row=period+1, column=day+1, value="数据处理错误")
            
            # 调整列宽和行高
            for col in range(1, 7):
                sheet.column_dimensions[get_column_letter(col)].width = 40
            for row in range(1, 7):
                sheet.row_dimensions[row].height = 40
            
            # 设置单元格对齐方式
            for row in sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=6):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    
    # 保存文件
    output_filename = 'empty_classrooms.xlsx'
    print(f"\n正在保存结果到 {output_filename}...")
    output_workbook.save(output_filename)
    print("导出完成！")

def main():
    print("Excel课程教室信息读取程序")
    print("请确保'input.xlsx'文件和classrooms.json在当前目录下")
    
    # 加载所有教室信息
    all_classrooms = load_classroom_info()
    
    while True:
        try:
            print("\n请选择操作：")
            print("1. 手动查询")
            print("2. 自动导出所有空闲教室")
            print("0. 退出程序")
            
            choice = int(input("请输入选项："))
            
            if choice == 0:
                break
            elif choice == 1:
                # 原有的手动查询代码
                print("\n请输入查询信息（输入0退出）：")
                day = int(input("请输入星期（1-5，周一到周五）："))
                if day == 0:
                    break
                
                if day < 1 or day > 5:
                    print("错误：星期必须在1-5之间")
                    continue
                
                period = int(input("请输入节次（1:1-2节, 2:3-4节, 3:5-6节, 4:7-8节, 5:9-10节）："))
                if period < 1 or period > 5:
                    print("错误：节次必须在1-5之间")
                    continue
                
                try:
                    row, col = get_row_col_by_time(day, period)
                    print(f"\n查询周{day} 第{format_time_period(period)}节的课程 (行{row}, 列{col})")
                except ValueError as e:
                    print(f"错误：{str(e)}")
                    continue
                
                results, classroom_infos = read_excel_cells('input1.xlsx', row, col)
                print("\n所有查询结果：")
                for result in results:
                    print(result)
                
                if classroom_infos and all_classrooms:  # 确保有教室信息和配置文件
                    while True:
                        try:
                            current_week = int(input("\n请输入当前周数（1-16，输入0返回）："))
                            if current_week == 0:
                                break
                            if current_week < 1 or current_week > 16:
                                print("错误：周数必须在1-16之间")
                                continue
                            
                            print(f"\n第{current_week}周 周{day} 第{format_time_period(period)}节的教室使用情况：")
                            used_classrooms = get_classrooms_by_week(classroom_infos, current_week)
                            if used_classrooms:
                                print(f"\n已使用教室（共{len(used_classrooms)}个）：")
                                for classroom in used_classrooms:
                                    print(classroom)
                                
                                # 获取并显示可用教室
                                available = get_available_classrooms(used_classrooms, all_classrooms)
                                if available:
                                    print(f"\n可用教室（共{len(available)}个）：")
                                    print("教室号\t\t类型\t\t容量")
                                    print("-" * 50)
                                    for room in available:
                                        print(f"{room['room_number']:<10}\t{room['room_type']:<10}\t{room['capacity']}人")
                                else:
                                    print("\n没有可用教室")
                            else:
                                print("本时段没有课程，所有教室都可用：")
                                print("教室号\t\t类型\t\t容量")
                                print("-" * 50)
                                for room_number, info in all_classrooms.items():
                                    print(f"{room_number:<10}\t{info['room_type']:<10}\t{info['capacity']}人")
                                
                        except ValueError:
                            print("错误：请输入有效的数字")
            elif choice == 2:
                export_to_excel(all_classrooms)
            else:
                print("无效的选项，请重新输入")
            
        except ValueError:
            print("错误：请输入有效的数字")
        
        print("\n")

if __name__ == "__main__":
    main()
