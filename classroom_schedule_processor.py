import openpyxl
import csv
import re
import os  # 添加 os 模块用于文件操作
from datetime import datetime  # 添加 datetime 用于生成时间戳
from itertools import islice
from tqdm import tqdm
from openpyxl.styles import Alignment, Font  # 导入 Font
from openpyxl.utils import get_column_letter

# --- Constants ---
SOURCE_DIR = "source"
OUTPUT_DIR = "."
CAMPUS_MAP = {
    "TianXin": "天心校区",
    "XiaoXiang": "潇湘校区",
    "XingLin": "杏林校区",
    "YueLuShan": "岳麓山校区"
}
MAX_WEEKS = 16 # Assume 16 weeks

def find_file_pairs(directory=SOURCE_DIR):
    """在指定目录中查找成对的 .csv 和 .xlsx 文件"""
    pairs = {}
    try:
        files = os.listdir(directory)
        csv_files = {os.path.splitext(f)[0] for f in files if f.lower().endswith('.csv')}
        xlsx_files = {os.path.splitext(f)[0] for f in files if f.lower().endswith('.xlsx')}
        
        common_basenames = sorted(list(csv_files.intersection(xlsx_files)))
        
        if not common_basenames:
            print(f"在 '{directory}' 目录下未找到匹配的 .csv 和 .xlsx 文件对。")
            return {}
            
        print(f"在 '{directory}' 目录下找到以下文件对:")
        pairs_dict = {}
        for i, basename in enumerate(common_basenames):
            pairs_dict[i + 1] = basename
            print(f"  {i + 1}: {basename}")
        return pairs_dict
        
    except FileNotFoundError:
        print(f"错误：目录 '{directory}' 不存在。")
        return {}
    except Exception as e:
        print(f"查找文件对时出错：{str(e)}")
        return {}

def load_classrooms(csv_filepath):
    """从指定的CSV文件加载教室列表"""
    classrooms = set()
    try:
        # 尝试用 gbk 解码，如果失败则尝试 utf-8
        try:
            with open(csv_filepath, 'r', encoding='gbk') as f:
                reader = csv.reader(f)
                for row in reader:
                    if row:
                        # Strip potential whitespace from classroom names
                        cleaned_name = row[0].strip()
                        if cleaned_name:
                             classrooms.add(cleaned_name)
        except UnicodeDecodeError:
            print(f"警告：使用 gbk 解码 '{csv_filepath}' 失败，尝试使用 utf-8...")
            with open(csv_filepath, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    if row:
                        cleaned_name = row[0].strip()
                        if cleaned_name:
                            classrooms.add(cleaned_name)
        return classrooms
    except FileNotFoundError:
        print(f"错误：CSV文件 '{csv_filepath}' 未找到。")
        return set()
    except Exception as e:
        print(f"加载教室列表 '{csv_filepath}' 时出错：{str(e)}")
        return set()

def parse_time_periods(time_str):
    """解析时间字符串，返回(星期几, 节次列表)的元组"""
    if not time_str or not isinstance(time_str, str):
        return None
    
    try:
        # 提取星期几（第一个数字）
        day = int(time_str[0])
        if day < 1 or day > 7: # Validate day is 1-7
             return None
        if day == 6 or day == 7:  # 忽略周六周日
            return None
        
        # 提取课节
        periods = []
        for i in range(1, len(time_str), 2):
            if i + 1 < len(time_str):
                period_str = time_str[i:i+2]
                if period_str.isdigit():
                    period = int(period_str)
                    if 1 <= period <= 10:
                        period_group = (period + 1) // 2  # 将1-10转换为1-5
                        if period_group not in periods:
                            periods.append(period_group)
        
        return (day, sorted(periods)) if periods else None
    except (ValueError, IndexError):
         # Handle cases where time_str is not in the expected format
         # print(f"警告: 无法解析时间字符串 '{time_str}'")
         return None

def parse_weeks(week_range, single_double=None):
    """解析周次范围，返回周数集合"""
    weeks = set()
    
    # Handle potential float inputs from Excel
    if isinstance(week_range, float):
        week_range = str(int(week_range))
    if isinstance(single_double, float):
         single_double = str(int(single_double))

    # 如果week_range为空且single_double为空，返回空集合
    if not week_range and not single_double:
        return weeks
    
    # 如果week_range是空的但有single_double值，说明是单个周数
    # Check if single_double might represent the week directly
    if not week_range and single_double:
        try:
            # Handle cases like "10" or "单周", "双周"
            if single_double.isdigit():
                 week = int(single_double)
                 if 1 <= week <= MAX_WEEKS:
                     weeks.add(week)
                 return weeks
            # If single_double is not a digit (like "单周"), it doesn't represent a single week number here.
            # We proceed assuming week_range might still hold the number if single_double just has text.
        except ValueError:
             pass # Fall through if conversion fails

    # If week_range is not None or empty string, process it.
    if week_range:
         # 处理多个部分（用逗号分隔）
        parts = str(week_range).split(',')
        
        for part in parts:
            part = part.strip()
            if '-' in part:
                # 处理范围（如"1-8"）
                try:
                    start, end = map(int, part.split('-'))
                    start = max(1, start)
                    end = min(MAX_WEEKS, end)
                    valid_range = range(start, end + 1)
                    
                    if single_double == '单周':
                        weeks.update(w for w in valid_range if w % 2 == 1)
                    elif single_double == '双周':
                        weeks.update(w for w in valid_range if w % 2 == 0)
                    else:  # 全周 or single_double is None/empty/numeric
                        weeks.update(valid_range)
                except ValueError:
                    continue # Skip malformed ranges like "1-周"
            else:
                # 处理单个数字
                try:
                    week = int(part)
                    if 1 <= week <= MAX_WEEKS:
                        # Apply single/double week logic if applicable *here*
                        is_odd = week % 2 == 1
                        if single_double == '单周' and is_odd:
                             weeks.add(week)
                        elif single_double == '双周' and not is_odd:
                             weeks.add(week)
                        elif single_double != '单周' and single_double != '双周': # Includes None, empty, numeric, or other text
                             weeks.add(week)
                except ValueError:
                    continue # Skip non-numeric parts
    
    # Final check if weeks is empty and single_double is a valid week number
    # This handles cases where week_range was something like "单周" and single_double had the week num
    if not weeks and single_double and single_double.isdigit():
         try:
            week = int(single_double)
            if 1 <= week <= MAX_WEEKS:
                weeks.add(week)
         except ValueError:
            pass
            
    return weeks

def load_schedule_data(basename, source_dir=SOURCE_DIR):
    """加载单个校区的教室和完整课程表数据"""
    csv_filepath = os.path.join(source_dir, f"{basename}.csv")
    xlsx_filepath = os.path.join(source_dir, f"{basename}.xlsx")
    print(f"-- 正在加载 {basename} 数据... --")
    
    all_classrooms = load_classrooms(csv_filepath)
    if not all_classrooms:
        print(f"错误：无法从 '{csv_filepath}' 加载教室列表，跳过 {basename}")
        return None, None

    used_classrooms = {} # {(week, day, period): {classrooms...}}
    workbook = None
    try:
        workbook = openpyxl.load_workbook(xlsx_filepath, read_only=True, data_only=True)
        sheet = workbook.active
        total_rows = sheet.max_row
        
        print(f"正在读取 {basename} 的课程信息 ({total_rows-3} 条)...")
        with tqdm(total=total_rows - 3, desc=f"处理 {basename} 数据", ncols=100, leave=False) as pbar:
            for row in sheet.iter_rows(min_row=4):
                try:
                    time_str = str(row[6].value).strip() if row[6].value else None
                    classroom_raw = str(row[7].value).strip() if row[7].value else None
                    week_range = str(row[8].value).strip() if row[8].value else None
                    single_double = str(row[9].value).strip() if row[9].value else None
                    
                    if time_str and classroom_raw:
                        time_info = parse_time_periods(time_str)
                        if time_info:
                            day, periods = time_info
                            weeks = parse_weeks(week_range, single_double)
                            
                            # Standardize classroom names found
                            found_classrooms_in_cell = re.findall(r'[A-Za-z]\d{2}\s\d{3}|[A-Za-z]\d+\s*\d*|[A-Za-z]{1,3}\d{1,3}', classroom_raw)
                            processed_classrooms = set()
                            if found_classrooms_in_cell:
                                processed_classrooms.update(c.strip() for c in found_classrooms_in_cell)
                            elif classroom_raw: # If regex didn't match but field is not empty
                                processed_classrooms.add(classroom_raw)

                            if not processed_classrooms:
                                continue # Skip if no classroom could be identified

                            for week in weeks:
                                if 1 <= week <= MAX_WEEKS:
                                    for period in periods:
                                        key = (week, day, period)
                                        if key not in used_classrooms:
                                            used_classrooms[key] = set()
                                        used_classrooms[key].update(processed_classrooms)
                
                except Exception as e:
                    # Avoid printing error for every row, maybe log it
                    # print(f"\n处理 {basename} 的行 {row[0].row} 时出错：{str(e)}") 
                    pass # Continue processing other rows
                
                pbar.update(1)
        print(f"{basename} 数据加载完成.")
        return all_classrooms, used_classrooms

    except FileNotFoundError:
        print(f"错误：文件 '{xlsx_filepath}' 未找到。")
        return None, None
    except Exception as e:
        print(f"加载 {basename} 的Excel文件时出错：{str(e)}")
        return None, None
    finally:
        if workbook:
            workbook.close()

def format_and_write_sheet(sheet, week, campus_name, status_text, all_classrooms, used_classrooms, title_font, maple_font):
     """格式化并填充单个工作表（用于特定周）"""
     # --- 添加标题行 (Row 1) ---
     sheet.insert_rows(1)
     sheet.merge_cells('A1:F1')
     title_cell = sheet['A1']
     title_string = f"第{week}周 {campus_name}{status_text}教室"
     title_cell.value = title_string
     title_cell.alignment = Alignment(horizontal='center', vertical='center')
     title_cell.font = title_font
     sheet.row_dimensions[1].height = 35
     # --- 标题行结束 ---

     # 设置表头 (Row 2 and Column 1 from Row 3 onwards)
     days = ["", "周一", "周二", "周三", "周四", "周五"]
     periods = ["", "1-2节", "3-4节", "5-6节", "7-8节", "9-10节"]

     # Day headers (Row 2)
     for col_idx, day in enumerate(days, 1):
         cell = sheet.cell(row=2, column=col_idx, value=day)
         cell.alignment = Alignment(horizontal='center', vertical='center')
         if col_idx > 1:
             cell.font = maple_font
     
     # Period headers (Column 1, starting Row 3)
     for row_idx, period_text in enumerate(periods, 1):
         if row_idx > 1:
             cell = sheet.cell(row=row_idx + 1, column=1, value=period_text)
             cell.alignment = Alignment(horizontal='center', vertical='center')
             cell.font = maple_font
     
     # 填充数据 (Starting Row 3, Column 2)
     data_start_row = 3
     data_start_col = 2
     show_occupied_flag = (status_text == "占用") # Determine based on status_text

     for day_idx in range(1, len(days)): # 1 to 5 (Mon to Fri)
         for period_idx in range(1, len(periods)): # 1 to 5 (1-2节 to 9-10节)
             key = (week, day_idx, period_idx)
             used = used_classrooms.get(key, set())
             
             available_classrooms = all_classrooms - used
             classrooms_to_show = used if show_occupied_flag else available_classrooms
             
             sorted_classrooms_raw = sorted(list(classrooms_to_show), key=lambda x: [int(t) if t.isdigit() else t.lower() for t in re.split('([0-9]+)', x)])
             
             modified_classrooms = []
             for name in sorted_classrooms_raw:
                 modified_name = name.replace('座', '')
                 modified_name = re.sub(r'(世[ABCD])(\d+)', r'\1 \2', modified_name)
                 modified_classrooms.append(modified_name)
                 
             cell_value = " ".join(modified_classrooms)
             target_row = period_idx + 1 + 1
             target_col = day_idx + 1
             data_cell = sheet.cell(row=target_row, column=target_col, value=cell_value)
             data_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
             data_cell.font = maple_font
     
     # 调整列宽
     sheet.column_dimensions[get_column_letter(1)].width = 15
     for col_letter_idx in range(data_start_col, data_start_col + len(days) - 1):
         sheet.column_dimensions[get_column_letter(col_letter_idx)].width = 50

def process_excel_all_weeks(basename, show_occupied, source_dir=SOURCE_DIR, output_dir=OUTPUT_DIR):
    """为单个校区生成包含所有周次的Excel文件"""
    campus_name = CAMPUS_MAP.get(basename, "未知校区")
    status_text = "占用" if show_occupied else "空闲"
    print(f"\n--- 开始为 {basename} ({campus_name}) 生成 {status_text} 教室表 (所有周次) ---")

    all_classrooms, used_classrooms = load_schedule_data(basename, source_dir)
    if all_classrooms is None or used_classrooms is None:
        print(f"无法加载 {basename} 的数据，跳过生成。")
        return
        
    output_workbook = openpyxl.Workbook()
    # Remove the default sheet created by Workbook()
    if len(output_workbook.sheetnames) > 0 and output_workbook.sheetnames[0] == 'Sheet':
         output_workbook.remove(output_workbook.active)

    maple_font = Font(name='Maple Mono Normal NL NF CN')
    title_font = Font(name='思源黑体 VF Medium', size=22)

    print(f"正在为 {basename} 生成 {MAX_WEEKS} 个周次的工作表...")
    for week in tqdm(range(1, MAX_WEEKS + 1), desc=f"生成 {basename} 表格", ncols=100):
        sheet_name = f"第{week}周"
        sheet = output_workbook.create_sheet(sheet_name)
        format_and_write_sheet(sheet, week, campus_name, status_text, all_classrooms, used_classrooms, title_font, maple_font)
        
    # 保存文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    status_filename = "occupied" if show_occupied else "empty"
    os.makedirs(output_dir, exist_ok=True)
    output_filename = os.path.join(output_dir, f'{basename}_{status_filename}_classrooms_{timestamp}.xlsx')
    
    print(f"\n正在保存 {basename} (所有周次) 的结果到 {output_filename}...")
    try:
        output_workbook.save(output_filename)
        print(f"{basename} (所有周次) 处理完成！")
    except Exception as e:
         print(f"保存文件 {output_filename} 时出错: {str(e)}")
    finally:
         if output_workbook:
             output_workbook.close()

def generate_single_week_excel(selected_week, show_occupied, source_dir=SOURCE_DIR, output_dir=OUTPUT_DIR):
    """为指定单周生成一个包含所有校区的整合Excel文件"""
    status_text = "占用" if show_occupied else "空闲"
    print(f"\n--- 开始为 第{selected_week}周 生成整合 {status_text} 教室表 --- ")

    available_pairs = find_file_pairs(source_dir)
    if not available_pairs:
        print("未找到任何校区文件对，无法生成整合文件。")
        return

    output_workbook = openpyxl.Workbook()
    if len(output_workbook.sheetnames) > 0 and output_workbook.sheetnames[0] == 'Sheet':
        output_workbook.remove(output_workbook.active)
        
    maple_font = Font(name='Maple Mono Normal NL NF CN')
    title_font = Font(name='思源黑体 VF Medium', size=22)

    for basename in available_pairs.values():
        campus_name = CAMPUS_MAP.get(basename, basename) # Use basename if not in map
        print(f"\n正在处理 {campus_name} (第{selected_week}周)..." )
        all_classrooms, used_classrooms = load_schedule_data(basename, source_dir)

        if all_classrooms is None or used_classrooms is None:
            print(f"无法加载 {basename} 的数据，跳过该校区。")
            continue

        sheet = output_workbook.create_sheet(campus_name) # Sheet name is campus name
        format_and_write_sheet(sheet, selected_week, campus_name, status_text, 
                               all_classrooms, used_classrooms, title_font, maple_font)
        print(f"{campus_name} (第{selected_week}周) 处理完成。")

    # 保存文件
    if not output_workbook.sheetnames: # Check if any sheets were actually created
         print("\n错误：未能成功处理任何校区的数据，未生成文件。")
         return
         
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    status_filename = "occupied" if show_occupied else "empty"
    os.makedirs(output_dir, exist_ok=True)
    output_filename = os.path.join(output_dir, f'Week_{selected_week}_{status_filename}_classrooms_{timestamp}.xlsx')
    
    print(f"\n正在保存 第{selected_week}周 (所有校区) 的整合结果到 {output_filename}...")
    try:
        output_workbook.save(output_filename)
        print(f"第{selected_week}周 整合文件处理完成！")
    except Exception as e:
         print(f"保存整合文件 {output_filename} 时出错: {str(e)}")
    finally:
         if output_workbook:
             output_workbook.close()

# --- Main execution block ---
if __name__ == "__main__":
    while True:
        print("\n╔══════════════════════════════════════╗")
        print("║       教室安排表生成器       ║")
        print("╠══════════════════════════════════════╣")
        print("║ 1. 生成 空闲 教室表                  ║")
        print("║ 2. 生成 占用 教室表                  ║")
        print("║ 0. 退出程序                          ║")
        print("╚══════════════════════════════════════╝")
        
        try:
            choice = input("请输入选项：").strip()
            if not choice.isdigit():
                 print("无效输入，请输入数字。")
                 continue
            choice = int(choice)
            
            if choice == 0:
                break
            elif choice in [1, 2]:
                show_occupied = (choice == 2)
                status_text_choice = "占用" if show_occupied else "空闲"
                
                # Ask for mode: All Weeks or Specific Week
                print("\n请选择生成模式:")
                print("  1. 所有周次 (每个校区一个文件，包含1-16周)")
                print("  2. 指定单周 (一个文件，每个校区一个Sheet，仅含指定周)")
                print("  0. 返回上级菜单")
                
                try:
                     mode_choice = input("请输入生成模式选项：").strip()
                     if not mode_choice.isdigit():
                         print("无效输入，请输入数字。")
                         continue
                     mode_choice = int(mode_choice)

                     if mode_choice == 0:
                         continue # Go back to main menu

                     elif mode_choice == 1: # All Weeks Mode
                        available_pairs = find_file_pairs(SOURCE_DIR)
                        if not available_pairs:
                            continue 
                            
                        print(f"\n请选择要处理的文件 ({status_text_choice}表 - 所有周次模式):" )
                        print("  0. 处理所有找到的文件对")
                        # File pairs already printed by find_file_pairs
                        
                        try:
                            file_choice_str = input("请输入文件选项 (输入数字): ").strip()
                            if not file_choice_str.isdigit():
                                print("无效输入，请输入数字。")
                                continue
                            file_choice = int(file_choice_str)
                            
                            if file_choice == 0:
                                print("\n将处理所有找到的文件对 (所有周次模式)...")
                                for basename in available_pairs.values():
                                    process_excel_all_weeks(basename, show_occupied=show_occupied, source_dir=SOURCE_DIR, output_dir=OUTPUT_DIR)
                                print("\n所有文件 (所有周次模式) 处理完成。")
                            elif file_choice in available_pairs:
                                selected_basename = available_pairs[file_choice]
                                process_excel_all_weeks(selected_basename, show_occupied=show_occupied, source_dir=SOURCE_DIR, output_dir=OUTPUT_DIR)
                            else:
                                print("无效的文件选项。")
                                
                        except ValueError:
                            print("错误：请输入有效的数字作为文件选项。")
                        except Exception as e:
                            print(f"处理 '所有周次' 文件选择时发生错误: {str(e)}")

                     elif mode_choice == 2: # Specific Week Mode
                        try:
                            week_str = input(f"请输入要生成的周次 (1-{MAX_WEEKS}): ").strip()
                            if not week_str.isdigit():
                                 print("无效输入，请输入数字。")
                                 continue
                            selected_week = int(week_str)
                            if not 1 <= selected_week <= MAX_WEEKS:
                                print(f"无效的周次，请输入 1 到 {MAX_WEEKS} 之间的数字。")
                                continue
                            
                            # Call the single week generation function (processes all campuses)
                            generate_single_week_excel(selected_week, show_occupied, source_dir=SOURCE_DIR, output_dir=OUTPUT_DIR)

                        except ValueError:
                            print("错误：请输入有效的数字作为周次。")
                        except Exception as e:
                             print(f"处理 '指定单周' 模式时发生错误: {str(e)}")
                     
                     else:
                          print("无效的生成模式选项。")

                except ValueError:
                     print("错误：请输入有效的数字作为模式选项。")
                except Exception as e:
                     print(f"处理生成模式选择时发生错误: {str(e)}")

            else:
                print("无效的操作选项，请重新输入。")
                
        except ValueError:
            print("错误：请输入有效的数字作为操作选项。")
        except Exception as e:
             print(f"主循环发生意外错误: {str(e)}")

    print("\n程序已退出。") 